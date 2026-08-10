"""
Excel report writer -- persists every pipeline run's incident reports into a
single accumulating .xlsx workbook.

Design (per project requirement): the agent talks to FOLDERS, and every run
APPENDS its reports to the SAME Excel file living in the same output folder,
so the workbook grows run-over-run as a cumulative audit/history -- it is never
overwritten. Two sheets:

  - "Incidents": one row per generated Report, fully flattened (severity,
    systems, components, time window, root-cause summary, recommended actions,
    evidence) plus a run_id + generated_at so you can tell which run each row
    came from.
  - "Run Log": one row per run with run-level metadata (mode, provider, file
    counts, events parsed/flagged, report count) -- the high-level "what
    happened in this run" companion to the per-incident detail.

Dependencies: openpyxl (added to requirements.txt). The rest of the pipeline
stays stdlib-only at its deterministic core; this module is only ever imported
by the entry points (main.py / pipeline_runner.py / app.py) that already pull
in the LLM stack.
"""

from __future__ import annotations

import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

INCIDENTS_SHEET = "Incidents"
RUN_LOG_SHEET = "Run Log"

INCIDENTS_HEADERS = [
    "run_id", "generated_at", "incident_id", "title", "severity",
    "systems_involved", "affected_components", "time_window", "confidence",
    "root_cause_summary", "recommended_actions", "evidence",
]

RUN_LOG_HEADERS = [
    "run_id", "generated_at", "mode", "provider", "mapping_csv",
    "zc_files", "es_files", "ats_files",
    "events_parsed", "events_flagged", "reports_count", "excel_path",
]

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")


def new_run_id() -> str:
    """A short, sortable-ish run identifier. Combined with generated_at this
    uniquely labels every row a run produces."""
    return f"run_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:6]}"


def _ensure_sheet(wb, name: str, headers: list[str]):
    """Return the worksheet for `name`, creating it with a styled header row
    if it doesn't exist yet (first run on a fresh workbook)."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    return ws


def _autofit(ws, min_width: int = 10, max_width: int = 80) -> None:
    """Best-effort column sizing so the workbook is readable on first open.
    Capped so a long root-cause paragraph doesn't make the column unusably
    wide -- Excel will still show the full content in the cell/editor."""
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=min_width,
        )
        ws.column_dimensions[letter].width = max(min_width, min(length + 2, max_width))


def _flatten_report(report, run_id: str, generated_at: str) -> list[Any]:
    """Turn one models.Report into a single spreadsheet row."""
    return [
        run_id,
        generated_at,
        report.incident_id,
        report.title,
        report.severity.value if hasattr(report.severity, "value") else str(report.severity),
        ", ".join(report.systems_involved),
        ", ".join(report.affected_components),
        report.time_window,
        report.confidence,
        report.root_cause_summary,
        "\n".join(f"- {a}" for a in report.recommended_actions),
        "\n".join(report.evidence),
    ]


def append_run(
    excel_path: str | Path,
    reports: list,
    run_meta: dict,
) -> str:
    """Append one run's reports + run-level metadata to the workbook at
    `excel_path`, creating it (with header rows) on first use.

    `reports`: list of models.Report (or any object with the same fields).
    `run_meta`: dict with keys mode, provider, mapping_csv, zc_files,
        es_files, ats_files, events_parsed, events_flagged.

    Returns the run_id that was stamped onto every row this run added, so
    callers can report it back to the user / log it.

    Never raises on an empty run -- an end-of-day batch that found nothing
    is still a real event worth a (zero-report) Run Log row, so the history
    shows that a run happened and produced nothing, rather than silently
    skipping the file.
    """
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    run_id = run_meta.get("run_id") or new_run_id()
    generated_at = run_meta.get("generated_at") or datetime.utcnow().isoformat()

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # Workbook() opens with a default empty "Sheet" -- drop it; our
        # _ensure_sheet calls create the real named sheets below.
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

    inc_ws = _ensure_sheet(wb, INCIDENTS_SHEET, INCIDENTS_HEADERS)
    run_ws = _ensure_sheet(wb, RUN_LOG_SHEET, RUN_LOG_HEADERS)

    for r in reports:
        inc_ws.append(_flatten_report(r, run_id, generated_at))

    run_ws.append([
        run_id,
        generated_at,
        run_meta.get("mode", ""),
        run_meta.get("provider", ""),
        run_meta.get("mapping_csv", ""),
        run_meta.get("zc_files", 0),
        run_meta.get("es_files", 0),
        run_meta.get("ats_files", 0),
        run_meta.get("events_parsed", 0),
        run_meta.get("events_flagged", 0),
        len(reports),
        str(excel_path),
    ])

    _autofit(inc_ws)
    _autofit(run_ws)

    # Freeze the header row so scrolling a long history stays usable.
    inc_ws.freeze_panes = "A2"
    run_ws.freeze_panes = "A2"

    wb.save(excel_path)
    return run_id


def report_count(excel_path: str | Path) -> int:
    """How many incident rows are already in the workbook (0 if it doesn't
    exist yet). Used by the UI to show 'appending to N existing rows'."""
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return 0
    try:
        wb = load_workbook(excel_path, read_only=True)
    except Exception:
        return 0
    if INCIDENTS_SHEET not in wb.sheetnames:
        return 0
    # max_row includes the header row -> subtract 1
    return max(wb[INCIDENTS_SHEET].max_row - 1, 0)